import pandas as pd
from openpyxl import load_workbook
import datetime,json, os, re

def extraire_premier_nombre_et_reste(chaine):
    pattern = r'(\d+)(.*)'  # Expression régulière pour capturer le premier nombre et le reste de la chaîne
    match = re.match(pattern, chaine)
    if match:
        nombre = match.group(1)  # Récupère le premier nombre trouvé
        reste = match.group(2)  # Récupère le reste de la chaîne
        try :
            return int(nombre), reste
        except TypeError as e:
            return int(nombre),None
    else:
        return "0", None  # Retourne None si aucun nombre n'est trouvé
    
def ExcelToPython(Variables):   
    #Variables 
    Nom_Feuille = Variables["Nom_Feuille"]
    NuméroLigneEnTetes = Variables["NuméroLigneEnTetes"]
    NomC_Entreprise = Variables["NomC_Entreprise"]
    NomC_BornesSimpl = Variables["NomC_BornesSimpl"]
    NomC_BornesDoubles = Variables["NomC_BornesDoubles"]
    NomC_Armoires = Variables["NomC_Armoires"]
    NomC_NbDemiJournees = Variables["NomC_NbDemiJournees"]
    NomC_Adresse = Variables["NomC_Adresse"]
    NomC_Departement = Variables["NomC_Departement"]
    NomC_Contact = Variables["NomC_Contact"]
    NomC_DateMiseEnService = Variables["NomC_DateMiseEnService"]
    NomC_Periodicite = Variables["NomC_Periodicite"]
    NomC_VisiteOrganisee = Variables["NomC_VisiteOrganisee"]
    path_excel = Variables["path_excel"]
    nom_fichierj = Variables["nom_fichierj"]

    #
    Contact=[]
    # Obtenez le répertoire de travail actuel
    repertoire_actuel = os.path.dirname(os.path.abspath(__file__))

    # Load the Excel workbook
    wb = load_workbook(path_excel)

    # Select the worksheet
    feuille = wb[Nom_Feuille]

    # Create an empty list to store the evaluated values
    data = []

    # Iterate over rows and columns to get evaluated values
    for row in feuille.iter_rows(values_only=True):
        row_values = []
        for cell in row:
            evaluated_value = cell
            row_values.append(evaluated_value)
        data.append(row_values)

    # Create a Pandas DataFrame from the data
    df = pd.DataFrame(data)
    nb_lignes, nb_colonnes = df.shape

    #Initialisation des Variables des numéros de colonnes
    numC_Entreprise=-1
    numC_Adresse=-1
    numC_Departement=-1
    numC_BornesSimpl=-1
    numC_BornesDoubles=-1
    numC_Armoires=-1
    numC_=-1
    numC_Contact=-1
    numC_DateMiseEnService=-1
    numC_Periodicite=-1
    numC_VisiteOrganisee = []

    #Parcourt la ligne des en-têtes pour trouver les colonnes relatives aux informations souhaitées
    for i in range(0,nb_colonnes) :
        if df.iloc[NuméroLigneEnTetes,i] is not None:
                if df.iloc[NuméroLigneEnTetes,i].lower()==NomC_Entreprise.lower() : #teste la correpondance sans prendre en compte les majuscules
                    numC_Entreprise = i #stocke le numéro de la colonne
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_BornesSimpl.lower() :
                    numC_BornesSimpl=i
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_BornesDoubles.lower() :
                    numC_BornesDoubles=i
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_Armoires.lower() :
                    numC_Armoires=i
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_NbDemiJournees.lower() :
                    numC_NbDemiJournes=i
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_Adresse.lower() :
                    numC_Adresse = i
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_Departement.lower() :
                    numC_Departement=i
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_Contact.lower() :
                    numC_Contact = i
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_DateMiseEnService.lower() :
                    numC_DateMiseEnService = i
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_Periodicite.lower() :
                    numC_Periodicite = i
                elif df.iloc[NuméroLigneEnTetes,i].lower()==NomC_VisiteOrganisee.lower() :
                    numC_VisiteOrganisee.append(str(i))#stocke chaque numéro de colonne relevé
    #Création des listes et dictionnaires globaux
    EnvoiJSON = {}
    date = str(datetime.datetime.today().strftime("%d/%m/%Y"))#date d'aujourd'hui convertie en string
    EnvoiJSON["dateDonnees"] = date
    ListesParcs = []
    
    for i in range(NuméroLigneEnTetes +1,nb_lignes):
        #Initialisation des Variables de données
        NbVisitesOrganisees = 0

        #Variables pour la lisibilité et traitement
        Entreprise = df.iloc[i,numC_Entreprise]
        BornesSimpl = extraire_premier_nombre_et_reste(str(df.iloc[i,numC_BornesSimpl]))
        BornesDoubles = extraire_premier_nombre_et_reste(str(df.iloc[i,numC_BornesDoubles]))
        Armoires = extraire_premier_nombre_et_reste(str(df.iloc[i,numC_Armoires]))
        


        NbDemiJournees = df.iloc[i,numC_NbDemiJournes]

        Adresse = df.iloc[i,numC_Adresse]
        Departement = df.iloc[i,numC_Departement]
        try:
            Contact = df.iloc[i,numC_Contact].split()
        except AttributeError as e :
            Contact=[]
        DateMiseEnService = df.iloc[i,numC_DateMiseEnService]
        Periodicite = df.iloc[i,numC_Periodicite]
        for j in range(0,len(numC_VisiteOrganisee)-1):
            if df.iloc[i,int(numC_VisiteOrganisee[j])] is not None:
                if df.iloc[i,int(numC_VisiteOrganisee[j])].lower()=="ok":
                    NbVisitesOrganisees +=1#le nombre de visites organisées depuis la mise en service = le nombre de périodes passées


        #Création des sous-sous-dictionnaires
        DictionnaireContact = {}
        if len(Contact)>0 :
            DictionnaireContact["prenom"] = Contact[0]
            if len(Contact)>1 :
                DictionnaireContact["nom"] = Contact[1]
                if len(Contact)>2 :
                        DictionnaireContact["telephone"] = Contact[2]
                        if len(Contact)>3 :
                            DictionnaireContact["email"] = Contact[3]


        DictionnaireBorneSimple = {}
        if(len(BornesSimpl)>0):
            DictionnaireBorneSimple["nb"] = BornesSimpl[0]
            if(len(BornesSimpl)>1):
             DictionnaireBorneSimple["type"] = BornesSimpl[1]
   
        DictionnaireBorneDouble = {}
        if(len(BornesDoubles)>0):
            DictionnaireBorneDouble["nb"] = BornesDoubles[0]
            if(len(BornesDoubles)>1):
                DictionnaireBorneDouble["type"] = BornesDoubles[1]
        
        DictionnaireArmoires = {}
        if(len(Armoires)>0):
            DictionnaireArmoires["nb"] = Armoires[0]
            if(len(Armoires)>1):
                DictionnaireArmoires["type"] = Armoires[1]

            DictionnaireMateriel = {}
            DictionnaireMateriel["borneSimple"]=DictionnaireBorneSimple
            DictionnaireMateriel["borneDouble"]=DictionnaireBorneDouble
            DictionnaireMateriel["armoire"]=DictionnaireArmoires
            #Dictionnaire Parc
            DictionnaireParc = {}
            DictionnaireParc["departement"] = Departement
            try :
                DictionnaireParc["dateMiseEnService"] = str(DateMiseEnService.strftime("%d/%m/%Y"))
            except AttributeError as e:
                DictionnaireParc["dateMiseEnService"] = None
            DictionnaireParc["periodiciteEnMois"] = Periodicite
            DictionnaireParc["nbVisitesOrganisees"] = NbVisitesOrganisees
            DictionnaireParc["nbDemiJoursTravail"] = NbDemiJournees
            DictionnaireParc["contact"] = DictionnaireContact
            DictionnaireParc["materiel"] = DictionnaireMateriel
            DictionnaireParc["nomEntreprise"] = str(Entreprise)
            DictionnaireParc["adresse"] = Adresse
            ListesParcs.append(DictionnaireParc)#On ajoute le dictionnaire du parc en cours à la liste des parcs

    EnvoiJSON["parcs"] = ListesParcs #On ajoute les parcs au Json 
    # Créez le chemin relatif en combinant le répertoire de travail et le nom du fichier
    chemin_fichier = os.path.join(repertoire_actuel, nom_fichierj)
    out_file = open(chemin_fichier, "w")
    json.dump(EnvoiJSON,out_file)#On enregistre le fichier Json
    out_file.close()
    return None